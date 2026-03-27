"""검색 결과 출력 포맷터 - 터미널, CSV, JSON, Markdown, Excel"""

import csv
import json
import io
from datetime import datetime
from pathlib import Path
from dataclasses import asdict

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich.markdown import Markdown

console = Console()


def _prepare_output_path(filepath: str) -> Path:
    """저장 경로의 부모 디렉토리를 미리 생성."""
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def print_results(articles: list, query: str = ""):
    """터미널에 결과를 예쁘게 출력."""
    if not articles:
        console.print("[yellow]검색 결과가 없습니다.[/yellow]")
        return

    success = [a for a in articles if a.get("success", True)]
    failed = [a for a in articles if not a.get("success", True)]

    console.print()
    console.print(Panel(
        f"[bold]검색어:[/bold] {query}\n"
        f"[bold]결과:[/bold] {len(articles)}건  "
        f"[green]성공 {len(success)}[/green]  "
        f"[red]실패 {len(failed)}[/red]  "
        f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M')}[/dim]",
        title="[bold blue]뉴스 수집 결과[/bold blue]",
        border_style="blue",
    ))

    for i, a in enumerate(articles, 1):
        title = a.get("title", "(제목 없음)")
        url = a.get("url", "")
        content = a.get("content", "")
        source = a.get("source", "")
        method = a.get("method", "")
        length = a.get("content_length", len(content))
        ok = a.get("success", True)

        status = "[green]OK[/green]" if ok else "[red]FAIL[/red]"
        source_tag = f" [dim]({source})[/dim]" if source else ""
        method_tag = f" [dim][{method}][/dim]" if method else ""

        console.print(f"\n  {status} [bold]{i}. {title}[/bold]{source_tag}{method_tag}")
        console.print(f"     [link={url}]{url}[/link]")

        if ok and content:
            preview = content[:200].replace("\n", " ")
            if len(content) > 200:
                preview += "..."
            console.print(f"     [dim]{preview}[/dim]")
            console.print(f"     [cyan]{length:,}자[/cyan]")
        elif not ok:
            error = a.get("error", "")
            if error:
                console.print(f"     [red]{error}[/red]")

    console.print()


def to_csv(articles: list, filepath: str | None = None, query: str = "") -> str:
    """CSV 형식으로 변환. filepath가 있으면 파일로 저장."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["키워드", "제목", "출처", "URL", "본문길이", "성공", "본문"])

    for i, a in enumerate(articles, 1):
        writer.writerow([
            a.get("keyword", query),
            a.get("title", ""),
            a.get("source", ""),
            a.get("url", ""),
            a.get("content_length", 0),
            "O" if a.get("success", True) else "X",
            a.get("content", ""),
        ])

    csv_text = output.getvalue()

    if filepath:
        path = _prepare_output_path(filepath)
        # newline="" 으로 Windows에서 빈 줄 방지
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            f.write(csv_text)
        console.print(f"[green]CSV 저장: {filepath}[/green]")

    return csv_text


def to_txt(articles: list, filepath: str, query: str = ""):
    """전체 정보 텍스트 파일로 저장."""
    path = _prepare_output_path(filepath)
    lines = []
    lines.append(f"뉴스 수집 결과: {query}")
    lines.append(f"수집일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"총 {len(articles)}건")
    lines.append("=" * 70)

    for i, a in enumerate(articles, 1):
        title = a.get("title", "")
        source = a.get("source", "")
        url = a.get("url", "")
        content = a.get("content", "")
        length = a.get("content_length", len(content))
        ok = a.get("success", True)

        lines.append(f"\n[{i}] {title}")
        lines.append(f"    출처: {source}")
        lines.append(f"    URL:  {url}")
        lines.append(f"    글자수: {length:,}자  {'성공' if ok else '실패'}")
        lines.append("-" * 70)
        lines.append(content)
        lines.append("=" * 70)

    text = "\n".join(lines) + "\n"
    path.write_text(text, encoding="utf-8")
    console.print(f"[green]TXT 저장: {filepath}[/green]")


def to_json(articles: list, filepath: str | None = None) -> str:
    """JSON 형식으로 변환."""
    data = {
        "collected_at": datetime.now().isoformat(),
        "count": len(articles),
        "articles": articles,
    }
    json_text = json.dumps(data, ensure_ascii=False, indent=2)

    if filepath:
        path = _prepare_output_path(filepath)
        path.write_text(json_text, encoding="utf-8")
        console.print(f"[green]JSON 저장: {filepath}[/green]")

    return json_text


def to_markdown(articles: list, query: str = "", filepath: str | None = None) -> str:
    """Markdown 형식으로 변환."""
    lines = [
        f"# 뉴스 수집: {query}",
        f"",
        f"> 수집일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}  ",
        f"> 결과: {len(articles)}건",
        "",
    ]

    for i, a in enumerate(articles, 1):
        title = a.get("title", "(제목 없음)")
        url = a.get("url", "")
        content = a.get("content", "")
        source = a.get("source", "")
        ok = a.get("success", True)

        lines.append(f"## {i}. {title}")
        lines.append(f"")
        lines.append(f"- 출처: {source}")
        lines.append(f"- URL: {url}")
        lines.append(f"- 상태: {'성공' if ok else '실패'}")
        lines.append(f"")

        if ok and content:
            lines.append(f"```")
            lines.append(content[:2000])
            if len(content) > 2000:
                lines.append(f"... (총 {len(content):,}자)")
            lines.append(f"```")
        lines.append("")

    md_text = "\n".join(lines)

    if filepath:
        path = _prepare_output_path(filepath)
        path.write_text(md_text, encoding="utf-8")
        console.print(f"[green]Markdown 저장: {filepath}[/green]")

    return md_text


def to_excel(articles: list, filepath: str, query: str = ""):
    """Excel(xlsx) 형식으로 저장. A1부터 바로 헤더+데이터."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "뉴스 수집 결과"

    # 스타일
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # A1부터 헤더
    headers = ["키워드", "제목", "출처", "URL", "본문길이", "성공", "본문"]
    col_widths = [12, 40, 15, 50, 10, 6, 80]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A" + chr(64 + col_idx - 26)].width = width

    ws.row_dimensions[1].height = 25

    # A2부터 데이터
    body_font = Font(name="맑은 고딕", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

    for i, a in enumerate(articles, 1):
        row = i + 1
        ok = a.get("success", True)
        fill = success_fill if ok else fail_fill

        values = [
            a.get("keyword", query),
            a.get("title", ""),
            a.get("source", ""),
            a.get("url", ""),
            a.get("content_length", 0),
            "O" if ok else "X",
            a.get("content", "")[:32000],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.fill = fill

    # 필터 + 틀 고정
    ws.auto_filter.ref = f"A1:I{len(articles) + 1}"
    ws.freeze_panes = "A2"

    # 저장
    path = _prepare_output_path(filepath)
    wb.save(str(path))
    console.print(f"[green]Excel 저장: {filepath}[/green]")
